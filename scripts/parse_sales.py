# -*- coding: utf-8 -*-
import openpyxl,datetime,json,io
wb=openpyxl.load_workbook('sales.xlsx',data_only=True)
def mon(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime('%b %Y')
    return '' if v is None else str(v)
def day(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime('%d %b %Y')
    return '' if v is None else str(v)
def n(v):
    if v is None: return None
    if isinstance(v,str): return v
    f=float(v); return int(f) if f==int(f) else round(f,2)

# ---- Monthly main (rows 2..12) ----
ms=wb['Monthly Summary']
monthly=[]
for r in range(2,13):
    a=ms.cell(r,1).value
    if a is None: continue
    row=[mon(a)]+[n(ms.cell(r,c).value) for c in range(2,14)]
    monthly.append(row)

# ---- No-GLP block ----
hdr=[mon(ms.cell(19,c).value) for c in range(2,8)]  # months
def block(r0,r1):
    out=[]
    for r in range(r0,r1+1):
        lab=ms.cell(r,1).value
        if lab is None: continue
        out.append([str(lab).strip()]+[n(ms.cell(r,c).value) for c in range(2,8)])
    return out
noglp={'months':hdr,'summary':block(20,22),'patients':block(25,33),'value':block(36,45),
       'needs':str(ms.cell(48,1).value)}

# ---- Day Level (2..29) ----
dl=wb['Day Level Breakdown']
days=[]
import re as _re
for r in range(2,dl.max_row+1):
    a=dl.cell(r,1).value
    if a is None: continue
    if isinstance(a,str):
        s=a.strip()
        if s.upper()!='TOTAL' and not _re.match(r'^\d{1,2}-[A-Za-z]{3}-\d{4}$',s): continue  # skip footnote text
    days.append([day(a)]+[n(dl.cell(r,c).value) for c in range(2,14)])

# ---- Raw Data ----
rd=wb['Raw Data']
def phone(v):
    if v is None: return ''
    if isinstance(v,float): return str(int(v))
    return str(v).split('.')[0]
raw=[]
for r in range(2,rd.max_row+1):
    a=rd.cell(r,1).value
    if a is None: continue
    g=[rd.cell(r,c).value for c in range(1,15)]
    raw.append([mon(g[0]),day(g[1]),g[2] or '',g[3] or '',g[4] or '',phone(g[5]),(str(g[6]) if g[6] is not None else ''),g[7] or '',n(g[8]),g[9] or '',g[10] or '',g[11] or '',g[12] or '',g[13] or ''])

io.open('sales_data.js','w',encoding='utf-8').write(
 'window.S_MONTHLY='+json.dumps(monthly,ensure_ascii=False)+';\n'+
 'window.S_NOGLP='+json.dumps(noglp,ensure_ascii=False)+';\n'+
 'window.S_DAYS='+json.dumps(days,ensure_ascii=False)+';\n'+
 'window.S_RAW='+json.dumps(raw,ensure_ascii=False)+';')
print('monthly',len(monthly),'days',len(days),'raw',len(raw))
print('sample month',monthly[0])
print('sample raw',raw[0])
print('noglp months',noglp['months'])
