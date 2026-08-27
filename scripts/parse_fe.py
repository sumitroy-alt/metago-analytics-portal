# -*- coding: utf-8 -*-
import warnings; warnings.simplefilter('ignore')
import openpyxl,datetime,json,io
wb=openpyxl.load_workbook('fe.xlsx',data_only=True)
def n(v):
    if v is None: return None
    if isinstance(v,str): return v
    try:
        f=float(v); return int(f) if f==int(f) else round(f,4)
    except: return v
def dstr(v):
    if isinstance(v,(datetime.datetime,datetime.date)):
        try: return v.strftime('%Y-%m-%d')
        except: return ''
    return '' if v is None else v

fsf=wb['FE Snippet and Funnel']
# monthly table cols 1..17, rows 2..13
monthly=[]
for r in range(2,14):
    a=fsf.cell(r,1).value
    if a is None: continue
    monthly.append([n(fsf.cell(r,c).value) for c in range(1,18)])
# legend: long text cells in cols 19..40
legend=[]
for r in range(1,15):
    for c in range(19,41):
        v=fsf.cell(r,c).value
        if isinstance(v,str) and len(v.strip())>28 and v.strip() not in legend:
            legend.append(v.strip())

# Data: full rows (cols 1..59)
dat=wb['Data']
hdr=[ (dat.cell(1,c).value or ('col%d'%c)) for c in range(1,60)]
rows=[]
for r in range(2,dat.max_row+1):
    if dat.cell(r,1).value is None and dat.cell(r,3).value is None: continue
    row=[]
    for c in range(1,60):
        v=dat.cell(r,c).value
        if c==4 and v is not None:  # phone
            try: v=str(int(float(v)))
            except: v=str(v)
        elif c==1 and v is not None:  # customer_id big
            try: v=str(int(float(v)))
            except: v=str(v)
        else:
            v=dstr(v) if isinstance(v,(datetime.datetime,datetime.date)) else n(v)
        row.append('' if v is None else v)
    rows.append(row)

io.open('fe_data.js','w',encoding='utf-8').write(
 'window.FE_MONTHLY='+json.dumps(monthly,ensure_ascii=False)+';\n'+
 'window.FE_LEGEND='+json.dumps(legend,ensure_ascii=False)+';\n'+
 'window.FE_HDR='+json.dumps([str(x) for x in hdr],ensure_ascii=False)+';\n'+
 'window.FE_DATA='+json.dumps(rows,ensure_ascii=False)+';')
print('monthly',len(monthly),'legend',len(legend),'data rows',len(rows),'cols',len(hdr))
print('month sample',monthly[0])
print('data sample',rows[0][:16])
