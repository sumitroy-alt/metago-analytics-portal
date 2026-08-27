# -*- coding: utf-8 -*-
import openpyxl, json, io, re, warnings; warnings.simplefilter('ignore')
wb = openpyxl.load_workbook('ltv.xlsx', data_only=True)
ws = wb['LTV Opportunity']

def n(v):
    if v is None: return None
    if isinstance(v, str): return v
    f = float(v); return int(f) if f == int(f) else round(f, 4)
def pct(v):
    if v is None: return ''
    if isinstance(v, str): return v          # '—' or 'n/a'
    return str(round(float(v) * 100)) + '%'
def cell(r, c): return ws.cell(r, c).value

# ---- A. opportunity (metrics rows 3-11) ----
opp = [[cell(r, 1), n(cell(r, 2))] for r in range(3, 12)]
onboarded = n(cell(4, 2)); collected = n(cell(5, 2)); cac = n(cell(6, 2))
# scenarios rows 3-5 : [take-up fraction, new revenue, new LTV drug, new LTV prog]
scen = [[n(cell(r, 3)), n(cell(r, 4)), n(cell(r, 5)), n(cell(r, 7))] for r in range(3, 6)]

# ---- summary rows 14-20 ----
summ = []
for r in range(14, 21):
    lab = cell(r, 1)
    if lab is None: continue
    summ.append([lab, n(cell(r, 2)), n(cell(r, 3)), lab.strip().lower().startswith('total')])

# ---- detail tree (rows 23-51 for DRUG/DIAG/DEVICES; 54-58 programme) ----
CAT = {'DRUG': 'drug', 'DIAGNOSTIC': 'diag', 'DEVICES': 'dev', 'SPECIAL PROGRAMME (ALT)': 'prog'}
def node(r, lbl, lvl, cat):
    return {'lbl': lbl, 'lvl': lvl, 'cat': cat,
            'price': cell(r, 4) or '', 'dur': cell(r, 5) or '', 'vpc': n(cell(r, 6)),
            'need': n(cell(r, 7)), 'pb': pct(cell(r, 8)),
            'cov': ('' if cell(r, 9) is None else n(cell(r, 9))), 'cp': pct(cell(r, 10)),
            'net': ('' if cell(r, 11) is None else n(cell(r, 11))),
            'gross': n(cell(r, 12)), 'nr': n(cell(r, 13)), 'children': []}

def build_tree(rows):
    roots = []; stack = []
    for r in rows:
        P, S, I = cell(r, 1), cell(r, 2), cell(r, 3)
        if P is None: continue
        if I not in (None, ''): lvl, lbl = 2, I
        elif S not in (None, ''): lvl, lbl = 1, S
        else: lvl, lbl = 0, P
        nd = node(r, lbl, lvl, CAT.get(P, ''))
        while stack and stack[-1][0] >= lvl: stack.pop()
        if lvl == 0 or not stack: roots.append(nd); stack = [(lvl, nd)]
        else: stack[-1][1]['children'].append(nd); stack.append((lvl, nd))
    return roots

detail = build_tree(range(23, 52))                 # DRUG, DIAGNOSTIC, DEVICES
# programme: row 54 is the (Parent+Sub) 'Condition tracks'; make a lvl0 'SPECIAL PROGRAMME'
prog = node(54, 'SPECIAL PROGRAMME (alt)', 0, 'prog')
for r in range(55, 59):
    it = node(r, cell(r, 3), 1, 'prog'); prog['children'].append(it)
detail.append(prog)

grand = {'lbl': 'GRAND TOTAL — programme excluded', 'price': 'mixed', 'dur': 'mixed',
         'vpc': n(cell(52, 6)), 'need': n(cell(52, 7)), 'pb': pct(cell(52, 8)),
         'cov': '—', 'cp': 'n/a', 'net': '—', 'gross': n(cell(52, 12)), 'nr': n(cell(52, 13))}
prog_total = {'lbl': 'TOTAL — less disease-mgmt drugs, plus programme',
              'price': '', 'dur': '', 'vpc': None, 'need': None, 'pb': '', 'cov': '', 'cp': '',
              'net': '', 'gross': n(cell(59, 12)), 'nr': n(cell(59, 13))}

# ---- Dr Rx tabs ----
def parse_rx(sheet, ncols):
    ws2 = wb[sheet]
    starts = [c for c in range(1, ws2.max_column + 1) if ws2.cell(4, c).value]  # month header cols
    data = {}
    for c0 in starts:
        hdr = str(ws2.cell(4, c0).value)
        m = re.match(r'([\d-]+)\D+(\d+) patients\D+(\d+) Rx\D+(\d+) doctors', hdr)
        month = m.group(1); meta = dict(patients=int(m.group(2)), rx=int(m.group(3)), docs=int(m.group(4)))
        rows = []; allrow = None
        r = 6
        while r <= ws2.max_row:
            name = ws2.cell(r, 1).value
            if name is None: r += 1; continue
            vals = [n(ws2.cell(r, c0 + k).value) for k in range(ncols)]
            if str(name).strip().upper() == 'ALL DOCTORS': allrow = vals
            else: rows.append([str(name)] + vals)
            r += 1
        data[month] = {'patients': meta['patients'], 'rx': meta['rx'], 'docs': meta['docs'],
                       'rows': rows, 'all': allrow}
    return data

fresh = parse_rx('Dr Rx · Fresh', 4)
repeat = parse_rx('Dr Rx · Repeat', 8)

LTV = {'opp': opp, 'scenarios': scen, 'summary': summ, 'detail': detail,
       'grand': grand, 'progTotal': prog_total, 'cac': cac,
       'collected': collected, 'onboarded': onboarded, 'fresh': fresh, 'repeat': repeat}
io.open('ltv_data.js', 'w', encoding='utf-8').write('window.LTV=' + json.dumps(LTV, ensure_ascii=False) + ';')
print('opp', len(opp), 'scen', len(scen), 'summary', len(summ), 'detail tops', [d['lbl'] for d in detail])
print('cac', cac, 'collected', collected, 'onboarded', onboarded)
print('fresh months', list(fresh.keys()), 'repeat months', list(repeat.keys()))
print('drug children', len(detail[0]['children']), 'first item', detail[0]['children'][0]['lbl'])
