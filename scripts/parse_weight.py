# -*- coding: utf-8 -*-
import openpyxl,datetime,json,io
wb=openpyxl.load_workbook('weight.xlsx',data_only=True)

def pct(v):
    if v is None or v=='' : return None
    try: return round(float(v)*100,1)
    except: return None

# ---------- drilldown parser (indent-based) ----------
def parse_tree(ws,r0,r1):
    roots=[];stack=[]
    for r in range(r0,r1+1):
        a=ws.cell(r,1).value
        if a is None: continue
        s=str(a)
        if s.strip()=='' : continue
        users=ws.cell(r,2).value
        if users is None: continue
        try: users=int(users)
        except: continue
        ind=len(s)-len(s.lstrip(' '))
        lvl=min(3,ind//3)
        m=[pct(ws.cell(r,3+i).value) for i in range(12)]
        node={'lbl':s.strip(),'users':users,'m':m,'children':[]}
        while stack and stack[-1][0]>=lvl: stack.pop()
        if lvl==0 or not stack: roots.append(node);stack=[(lvl,node)]
        else: stack[-1][1]['children'].append(node);stack.append((lvl,node))
    return roots

wd=wb['Weight Dashboard']
# find section boundaries by scanning col1
secs={}
for r in range(1,wd.max_row+1):
    a=wd.cell(r,1).value
    if a is None: continue
    s=str(a).strip()
    if s.startswith('3.'): secs['med']=r
    elif s.startswith('4.'): secs['coach']=r
# medication: from secs['med']+2 (skip header 'Group') to secs['coach']-1
med=parse_tree(wd,secs['med']+2,secs['coach']-2)
coach=parse_tree(wd,secs['coach']+2,wd.max_row)
io.open('med_data.js','w',encoding='utf-8').write('window.MED_FULL='+json.dumps(med,ensure_ascii=False)+';')
io.open('coach_data.js','w',encoding='utf-8').write('window.COACH_FULL='+json.dumps(coach,ensure_ascii=False)+';')
print('med tops:',[(x['lbl'],x['users']) for x in med])
print('coach tops:',len(coach),[x['lbl'] for x in coach][:6])

# section 1 & 2
def rowvals(ws,r,isPct):
    out=[]
    for c in range(2,14):
        v=ws.cell(r,c).value
        if v is None: out.append('')
        elif isPct: out.append(str(round(float(v)*100,1))+'%')
        else: out.append(int(v) if float(v)==int(float(v)) else round(float(v),1))
    return out
MOM=[]
labels_pct={'Coverage %','Avg cumulative loss %','Median cumulative loss %','Best %','% of users >=5% lost','% of users >=10% lost','% of users >=15% lost'}
for r in range(6,16):
    lab=wd.cell(r,1).value
    if not lab: continue
    MOM.append([str(lab).strip(),rowvals(wd,r,str(lab).strip() in labels_pct)])
TODAY=[['Users currently at this month',rowvals(wd,20,False)]]
io.open('mom_data.js','w',encoding='utf-8').write('window.MOM_FULL='+json.dumps(MOM,ensure_ascii=False)+';window.TODAY_FULL='+json.dumps(TODAY,ensure_ascii=False)+';')
print('MOM rows:',len(MOM))

# ---------- AUG full ----------
aug=wb['Aug 2026']
hdr=4
def dstr(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime('%Y-%m-%d')
    return '' if v is None else str(v)
def numv(v):
    if v is None: return None
    try:
        f=float(v); return int(f) if f==int(f) else round(f,2)
    except: return None
rows=[]
for r in range(hdr+1,aug.max_row+1):
    name=aug.cell(r,1).value
    if name is None or str(name).strip()=='' : continue
    g=[aug.cell(r,c).value for c in range(1,25)]
    nm=str(g[0]).strip()
    if ' +' in nm: nm=nm.split(' +')[0].strip()
    ph=g[1]
    ph='' if ph is None else (str(int(ph)) if isinstance(ph,float) else str(ph).split('.')[0])
    row=[nm,ph,g[2] or '',numv(g[3]),g[4] or '',numv(g[5]),g[6] or '',g[7] or '',g[8] or '',g[9] or '',g[10] or '',dstr(g[11]),g[12] or '',numv(g[13]),numv(g[14]),numv(g[15]),dstr(g[16]),
      (None if g[17] is None else str(round(float(g[17])*100,1))+'%'),
      (None if g[18] is None else str(round(float(g[18])*100,1))+'%'),
      numv(g[19]),g[20] or '',g[21] or '',numv(g[22]),numv(g[23])]
    rows.append(row)
io.open('aug_data.js','w',encoding='utf-8').write('window.AUG_FULL='+json.dumps(rows,ensure_ascii=False)+';')
print('AUG rows:',len(rows))
print('sample row5:',rows[0])
from collections import Counter
print('perf:',dict(Counter(r[20] for r in rows)))
print('vsPeer sample:',[r[19] for r in rows[:5]])
