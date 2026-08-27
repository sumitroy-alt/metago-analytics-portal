# -*- coding: utf-8 -*-
import warnings; warnings.simplefilter('ignore')
import openpyxl,json,io
wb=openpyxl.load_workbook('ret.xlsx',data_only=True)

def blocks(nm):
    ws=wb[nm]; out=[]; r=1; mx=ws.max_row
    while r<=mx:
        a=ws.cell(r,1).value
        if a is None: r+=1; continue
        nxt=ws.cell(r+1,1).value if r+1<=mx else None
        if nxt in ('Acq Month','Medication'):
            title=str(a).strip(); key=str(nxt); r+=2
            rows=[]
            while r<=mx:
                m=ws.cell(r,1).value
                if m is None: break
                nval=ws.cell(r,2).value
                counts=[ws.cell(r,c).value for c in range(3,11)]
                pcts=[ws.cell(r,c).value for c in range(12,20)]
                rows.append([str(m),(int(nval) if isinstance(nval,(int,float)) else nval),
                             [('' if x is None else str(x)) for x in pcts],
                             [('' if x is None else str(x)) for x in counts]])
                stop=(str(m).strip()=='ALL'); r+=1
                if stop: break
            out.append({'title':title,'key':key,'rows':rows})
        else:
            r+=1
    return out

LENSES={}
for nm in ['All Lens Coverage','DC Retention','Prescription Retention','Payment Retention']:
    LENSES[nm]=blocks(nm)

bt=wb['Blood Test']
def btblock(r0,r1):
    out=[]
    for r in range(r0,r1+1):
        a=bt.cell(r,1).value
        if a is None: continue
        row=[str(a)]+[bt.cell(r,c).value for c in range(2,9)]
        row=[('' if x is None else (int(x) if isinstance(x,float) and x==int(x) else x)) for x in row]
        out.append(row)
    return out
BT={'overall':btblock(6,16),'byMed':btblock(20,26)}

io.open('ret_data.js','w',encoding='utf-8').write(
 'window.RET_LENSES='+json.dumps(LENSES,ensure_ascii=False)+';\n'+
 'window.RET_BT='+json.dumps(BT,ensure_ascii=False)+';')
print({k:len(v) for k,v in LENSES.items()})
for b in LENSES['DC Retention']: print('  block:',b['title'],'|',b['key'],'|',len(b['rows']),'rows')
